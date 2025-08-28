import streamlit as st
import pandas as pd
import re
import numpy as np  # add at top
from io import BytesIO
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter


# ───── Streamlit setup ────────────────────────────────────────────────────────
st.set_page_config(page_title="Visitor List Cleaner", layout="wide")
st.title("🇸🇬 CLARITY GATE – VISITOR DATA CLEANING & VALIDATION 🫧")


# ───── 1) Info Banner ──────────────────────────────────────────────────────────
st.info(
    """
    **Data Integrity Is Our Foundation**  
    At every step—from file upload to final report—we enforce strict validation to guarantee your visitor data is accurate, complete, and compliant.  
    Maintaining integrity not only expedites gate clearance, it protects our facilities and ensures we meet all regulatory requirements.
    """
)

# ───── 2) Why Data Integrity? ───────────────────────────────────────────────────
with st.expander("Why is Data Integrity Important?"):
    st.write(
        """
        - **Accuracy**: Correct visitor details reduce clearance delays.  
        - **Security**: Reliable ID checks prevent unauthorized access.  
        - **Compliance**: Audit-ready records ensure regulatory adherence.  
        - **Efficiency**: Trustworthy data powers faster reporting and analytics.
        """
    )

# ───── Download Sample Template ────────────────────────────────────────────────
# This reads the Excel you committed as sample_template.xlsx in your repo root
with open("sample_template.xlsx", "rb") as f:
    sample_bytes = f.read()
st.download_button(
    label="🌟 Download Template",
    data=sample_bytes,
    file_name="sample_template.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

# ───── 3) Uploader & Warning ───────────────────────────────────────────────────

st.markdown(
    """<div style='font-size:14px; font-weight:bold; color:#38761d;'>
    Please ensure your spreadsheet has no missing or malformed fields.<br>
    Columns E and F are not required to be filled in.
    </div>""",
    unsafe_allow_html=True
)

uploaded = st.file_uploader("📁 Upload file", type=["xlsx"])

# ───── 4) Estimate Clearance Date ───────────────────────────────────────────────
now = datetime.now(ZoneInfo("Asia/Singapore"))
formatted_now = now.strftime("%A %d %B, %I:%M%p").lstrip("0")
#st.markdown("### 🗓️ Estimate Clearance Date 🍍")


# The Today timestamp:
st.write("**Today is:**", formatted_now)

if st.button("▶️ Earliest clearance:"):
    if now.time() >= datetime.strptime("15:00", "%H:%M").time():
        effective_submission_date = now.date() + timedelta(days=1)
    else:
        effective_submission_date = now.date()

    while effective_submission_date.weekday() >= 5:
        effective_submission_date += timedelta(days=1)

    working_days_count = 0
    estimated_date = effective_submission_date
    while working_days_count < 2:
        estimated_date += timedelta(days=1)
        if estimated_date.weekday() < 5:
            working_days_count += 1

    clearance_date = estimated_date
    while clearance_date.weekday() >= 5:
        clearance_date += timedelta(days=1)

    formatted = f"{clearance_date:%A} {clearance_date.day} {clearance_date:%B}"
    st.success(f" **{formatted}**")

# ───── Helper Functions ────────────────────────────────────────────────────────

def smart_title_case(name):
    words = name.strip().split()
    cleaned = []
    for w in words:
        if len(w) <= 3 and w.isupper():  # treat short uppercase as acronyms
            cleaned.append(w)
        else:
            cleaned.append(w.capitalize())
    return " ".join(cleaned)

def nationality_group(row):
    nat = str(row["Nationality (Country Name)"]).strip().lower()
    pr  = str(row["PR"]).strip().lower()
    if nat == "singapore":
        return 1
    elif pr in ("yes","y","pr"):
        return 2
    elif nat == "malaysia":
        return 3
    elif nat == "india":
        return 4
    else:
        return 5

def split_name(full_name):
    s = str(full_name).strip()
    if " " in s:
        i = s.find(" ")
        return pd.Series([s[:i], s[i+1:]])
    return pd.Series([s, ""])

def clean_gender(g):
    v = str(g).strip().upper()
    if v == "M": return "Male"
    if v == "F": return "Female"
    if v in ("MALE","FEMALE"): return v.title()
    return v

def normalize_pr(value):
    val = str(value).strip().lower()
    if val in ("pr", "yes", "y"):
        return "PR"
    elif val in ("n", "no", "na", "", "nan"):
        return ""
    else:
        return val.upper() if val.isalpha() else val

def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    # keep first 13 cols & rename
    df = df.iloc[:, :13]
    df.columns = [
        "S/N","Vehicle Plate Number","Company Full Name","Full Name As Per NRIC",
        "First Name as per NRIC","Middle and Last Name as per NRIC","Identification Type",
        "IC (Last 3 digits and suffix) 123A","Work Permit Expiry Date",
        "Nationality (Country Name)","PR","Gender","Mobile Number",
    ]
    df = df.dropna(subset=df.columns[3:13], how="all")

    # normalize company # Capitalize each word + fix Pte Ltd formatting
    df["Company Full Name"] = (
    df["Company Full Name"]
      .astype(str)
      .apply(smart_title_case)
      .str.replace(r"\bPte\s+Ltd\b", "Pte Ltd", flags=re.IGNORECASE, regex=True)
    )

    # standardize nationality
    nat_map = {"chinese":"China","singaporean":"Singapore","malaysian":"Malaysia","indian":"India"}
    nat_map = {
    "chinese": "China",
    "singaporean": "Singapore",
    "malaysian": "Malaysia",
    "indian": "India",
    "bangladeshi": "Bangladesh",
    "british": "United Kingdom",
    "uk": "United Kingdom",
    "england": "United Kingdom",
    "us": "United States",
    "usa": "United States",
    "u.s.": "United States",
    "u.s.a.": "United States"
}

    df["Nationality (Country Name)"] = (
        df["Nationality (Country Name)"]
          .astype(str).str.strip().str.lower()
          .replace(nat_map, regex=False)
          .str.title()
    )

    # sort & serial
    df["SortGroup"] = df.apply(nationality_group, axis=1)
    df = (
        df.sort_values(
            ["Company Full Name","SortGroup","Nationality (Country Name)","Full Name As Per NRIC"],
            ignore_index=True
        )
        .drop(columns="SortGroup")
    )
    df["S/N"] = range(1, len(df) + 1)

    # 🔄 Apply updated PR normalization
    df["PR"] = df["PR"].apply(normalize_pr)

    # normalize ID type
    df["Identification Type"] = (
        df["Identification Type"]
          .astype(str).str.strip()
          .apply(lambda v: "FIN" if v.lower() == "fin" else v.upper())
    )


    # Clean vehicle plate numbers
    df["Vehicle Plate Number"] = (
    df["Vehicle Plate Number"]
      .astype(str)
      .str.strip()
      .str.upper()
      .replace({r"(?i)^nan$": "", r"(?i)^nil$": ""}, regex=True)
      .str.replace(r"[\/,]", ";", regex=True)
      .str.replace(r"\s*;\s*", ";", regex=True)
      .str.replace(r"\s+", "", regex=True)
    )

    # split names 
    #df["Full Name As Per NRIC"] = df["Full Name As Per NRIC"].astype(str).str.title() 
    #df[["First Name as per NRIC","Middle and Last Name as per NRIC"]] = ( 
    #    df["Full Name As Per NRIC"].apply(split_name) )

    # ---------------- Split Names (enhanced & safe) ----------------
    # Normalize spacing/casing first
    df["Full Name As Per NRIC"] = (
        df["Full Name As Per NRIC"]
          .astype(str)
          .str.replace(r"\s+", " ", regex=True)
          .str.strip()
          .str.title()
    )
    # Explicit regex to avoid pandas None-pattern TypeError
    parts = df["Full Name As Per NRIC"].str.split(r"\s+", n=1, expand=True)
    df["First Name as per NRIC"] = parts[0]
    # If there’s no second token, copy first name into F
    df["Middle and Last Name as per NRIC"] = parts[1].fillna(parts[0])
    # ---------------------------------------------------------------


    # swap IC/WP if needed
    iccol, wpcol = "IC (Last 3 digits and suffix) 123A", "Work Permit Expiry Date"
    if df[iccol].astype(str).str.contains("-", na=False).any():
        df[[iccol, wpcol]] = df[[wpcol, iccol]]
    df[iccol] = df[iccol].astype(str).str[-4:]

    # clean mobile
    def fix_mobile(x):
        d = re.sub(r"\D", "", str(x))
        if len(d) > 8:
            extra = len(d) - 8
            if d.endswith("0"*extra): d = d[:-extra]
            else: d = d[-8:]
        if len(d) < 8: d = d.zfill(8)
        return d
    df["Mobile Number"] = df["Mobile Number"].apply(fix_mobile)

    df["Gender"] = df["Gender"].apply(clean_gender)
    df[wpcol] = pd.to_datetime(df[wpcol], errors="coerce").dt.strftime("%Y-%m-%d")

    return df
    
def generate_visitor_only(df: pd.DataFrame) -> BytesIO:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Visitor List")
        ws = writer.sheets["Visitor List"]

        header_fill  = PatternFill("solid", fgColor="94B455")
        warning_fill = PatternFill("solid", fgColor="DA9694")
        border       = Border(*[Side("thin")]*4)
        center       = Alignment("center","center")
        normal_font  = Font(name="Calibri", size=9)
        bold_font    = Font(name="Calibri", size=9, bold=True)

        # style all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border    = border
                cell.alignment = center
                cell.font      = normal_font

        # header row
        for col in range(1, ws.max_column + 1):
            h = ws[f"{get_column_letter(col)}1"]
            h.fill = header_fill
            h.font = bold_font
        ws.freeze_panes = "B2"

        errors = 0
        seen = {}                    # ← initialize duplicate‐tracker

        for r in range(2, ws.max_row + 1):
            # pull values
            idt = str(ws[f"G{r}"].value).strip().upper()
            nat = str(ws[f"J{r}"].value).strip().title()
            pr  = str(ws[f"K{r}"].value).strip().lower()
            wpd = str(ws[f"I{r}"].value).strip()
            name = str(ws[f"D{r}"].value or "").strip()   # ← grab “Full Name” from col D

            bad = False


             ─── highlight if expiry date is expired OR within 6 months ───
            expiry_str = str(ws[f"I{r}"].value).strip()
            try:
                expiry_date = datetime.strptime(expiry_str, "%Y-%m-%d").date()
                today_sg = datetime.now(ZoneInfo("Asia/Singapore")).date()
                six_months_ahead = today_sg + timedelta(days=180)  # ≈ 6 months
            
                # Note: <= six_months_ahead already covers "expired" as well
                if expiry_date <= six_months_ahead:
                    # highlight just the expiry cell:
                    ws[f"I{r}"].fill = warning_fill
                    errors += 1
            
            
            except ValueError:
                # If invalid date, overwrite with "Invalid" and highlight
                ws[f"I{r}"].value = "Invalid"
                ws[f"I{r}"].fill = warning_fill
                errors += 1

            expiry_issue = False  # track if we've already flagged the expiry cell
            

            # ── NEW RULE: Singaporeans cannot be PR ────────────────────────────
            if nat == "Singapore" and pr == "pr":
                bad = True
            
            if idt != "NRIC" and pr == "pr": bad = True
            if idt == "FIN" and (nat == "Singapore" or pr == "pr"): bad = True
            if idt == "NRIC" and not (nat == "Singapore" or pr == "pr"): bad = True
            if idt == "FIN" and not wpd: bad = True
            if idt == "WP" and not wpd: bad = True

            if bad:
             # highlight the offending cells
                for col in ("G","J","K","I"):
                    ws[f"{col}{r}"].fill = warning_fill
                errors += 1

            # ─── duplicate‐check on column D ──────────────────────────
            if name:
                if name in seen:
                    # highlight both the old row and the new row
                    ws[f"D{r}"].fill = warning_fill
                    ws[f"D{seen[name]}"].fill = warning_fill
                    errors += 1
                else:
                    seen[name] = r

        if errors:
            st.warning(f"⚠️ {errors} validation error(s) found.")
        
        # Set fixed column widths
        column_widths = {
            "A": 3.38,
            "C": 23.06,
            "D": 17.25,
            "E": 17.63,
            "F": 26.25,
            "G": 13.94,
            "H": 24.06,
            "I": 18.38,
            "J": 20.31,
            "K": 4,
            "L": 5.81,
            "M": 11.5,
        }
        # B is dynamic (auto-fit based on max content)
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            if col_letter == "B":
                width = max(len(str(cell.value)) for cell in col if cell.value)
                ws.column_dimensions[col_letter].width = width
            elif col_letter in column_widths:
                ws.column_dimensions[col_letter].width = column_widths[col_letter]
        
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 16.8

        # vehicles summary
        plates = []
        for v in df["Vehicle Plate Number"].dropna():
            plates += [x.strip() for x in str(v).split(";") if x.strip()]
        ins = ws.max_row + 2
        if plates:
            ws[f"B{ins}"].value = "Vehicles"
            ws[f"B{ins}"].font = Font(size=9)
            ws[f"B{ins}"].border = border
            ws[f"B{ins}"].alignment = center

            ws[f"B{ins+1}"].value = ";".join(sorted(set(plates)))
            ws[f"B{ins+1}"].font = Font(size=9)
            ws[f"B{ins+1}"].border = border
            ws[f"B{ins+1}"].alignment = center
            ins += 2

        ws[f"B{ins}"].value = "Total Visitors"
        ws[f"B{ins}"].font = Font(size=9)
        ws[f"B{ins}"].border = border
        ws[f"B{ins}"].alignment = center

        ws[f"B{ins+1}"].value = df["Company Full Name"].notna().sum()
        ws[f"B{ins+1}"].font = Font(size=9)
        ws[f"B{ins+1}"].border = border
        ws[f"B{ins+1}"].alignment = center

    buf.seek(0)
    return buf

    
# ───── Read, Clean & Download ────────────────────────────────────────────────
if uploaded:
    raw_df = pd.read_excel(uploaded, sheet_name="Visitor List")

    
    company_cell = raw_df.iloc[0, 2]
    company = (
        str(company_cell).strip()
        if pd.notna(company_cell) and str(company_cell).strip()
        else "VisitorList"
    )

    cleaned = clean_data(raw_df)
    out_buf = generate_visitor_only(cleaned)

    today_str = datetime.now(ZoneInfo("Asia/Singapore")).strftime("%Y%m%d")
    fname = f"{company}_{today_str}.xlsx"

    st.download_button(
        label="📥 Download Cleaned Visitor List",
        data=out_buf.getvalue(),
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.caption(
        "✅ Your data has been validated. Please double-check critical fields before sharing with DC team."
    )


# ───── 5) Final Notice (always shown) ────────────────────────────────────────
st.markdown(
    """
    <div style="line-height:1.2; font-size:14px;">
      We will do our utmost to deliver your access ticket 1 day before your scheduled entry.<br>
      Kindly ensure that approved access clearance codes are obtained before planning or commencing any work activities in the data center.<br>
      Please be reminded to go through the Clarity Gate prior to submission, and ensure that all visitor and shipment details are complete and accurate to prevent rescheduling due to clarification.<br><br>
      <strong>Note:</strong><br>
      The Clarity Gate operates on the GOFAI system, which relies on explicitly programmed rules and logic.<br>
      Although its validation accuracy can reach up to 98%, we strongly recommend that you thoroughly review all information before submission.<br>
      Thank you for your cooperation.<br><br>
    </div>
    """,
    unsafe_allow_html=True,
)


# ───── 6) Vendor Accuracy Reminder (always shown) ────────────────────────────
st.markdown(
    """
    <div style="line-height:1.2; font-size:14px;">
      <strong>Kindly remind all vendors to take the accuracy of the submitted information seriously.</strong><br>
      Any <em>incorrect or incomplete details</em> will result in <em>rejection</em>, and the personnel will not be allowed to enter the data centre.<br>
      <em>This requirement is non-negotiable, and strict compliance is expected.<em><br>
      Please ensure this message is clearly conveyed to all concerned.
    </div>
    """,
    unsafe_allow_html=True,
)
